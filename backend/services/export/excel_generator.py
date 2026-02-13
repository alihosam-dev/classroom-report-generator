"""
Excel report generator
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
import os

class ExcelGenerator:
    def generate(self, course, students, coursework, submissions, output_path, include_grades=True):
        """Generate Excel file with raw grades and analysis"""
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create raw grades sheet
        self._create_raw_grades_sheet(wb, students, coursework, submissions, include_grades)
        
        # Create analysis sheet
        if include_grades:
            self._create_analysis_sheet(wb, students, coursework, submissions)
        
        # Ensure output directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # Save workbook
        wb.save(output_path)
        
        return output_path
    
    def _create_raw_grades_sheet(self, wb, students, coursework, submissions, include_grades):
        """Create sheet with raw grades"""
        ws = wb.create_sheet('Raw Grades')
        
        # Define styles
        header_fill = PatternFill(start_color='2E5090', end_color='2E5090', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        alt_row_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        
        # Headers
        headers = ['Student Name', 'Email'] + [cw['title'] for cw in coursework]
        if include_grades:
            headers.append('Average')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows
        for row_idx, student in enumerate(students, 2):
            student_name = student.get('profile', {}).get('name', {}).get('fullName', 'Unknown')
            student_email = student.get('profile', {}).get('emailAddress', '')
            student_id = student.get('userId')
            
            # Alternating row colors
            if row_idx % 2 == 0:
                row_fill = alt_row_fill
            else:
                row_fill = PatternFill()
            
            # Student info cells
            name_cell = ws.cell(row=row_idx, column=1, value=student_name)
            name_cell.font = Font(bold=True, size=10)
            name_cell.fill = row_fill
            name_cell.border = border
            
            email_cell = ws.cell(row=row_idx, column=2, value=student_email)
            email_cell.font = Font(size=10)
            email_cell.fill = row_fill
            email_cell.border = border
            
            grades = []
            max_points = []
            for col_idx, cw in enumerate(coursework, 3):
                cw_id = cw['id']
                cw_submissions = submissions.get(cw_id, [])
                max_pts = cw.get('maxPoints', 100)
                
                # Find this student's submission
                grade = None
                submission_state = None
                for sub in cw_submissions:
                    if sub.get('userId') == student_id:
                        grade = sub.get('assignedGrade')
                        submission_state = sub.get('state')
                        break
                
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = row_fill
                cell.border = border
                cell.font = Font(size=10)
                
                if include_grades and grade is not None:
                    cell.value = f"{grade}/{max_pts}"
                    grades.append(grade)
                    max_points.append(max_pts)
                elif submission_state == 'TURNED_IN' or submission_state == 'RETURNED':
                    cell.value = 'Not Graded Yet'
                    cell.font = Font(italic=True, color='FF8C00', size=10)
                else:
                    cell.value = 'Not Submitted'
                    cell.font = Font(italic=True, color='DC143C', size=10)
            
            # Calculate average as percentage
            if include_grades and grades:
                total_earned = sum(grades)
                total_possible = sum(max_points)
                avg_pct = (total_earned / total_possible * 100) if total_possible > 0 else 0
                avg_cell = ws.cell(row=row_idx, column=len(coursework) + 3, value=f"{round(avg_pct, 1)}%")
                avg_cell.font = Font(bold=True, size=10)
                avg_cell.alignment = Alignment(horizontal='center', vertical='center')
                avg_cell.fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
                avg_cell.border = border
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    # Skip merged cells
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                except:
                    pass
            if column_letter:
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_analysis_sheet(self, wb, students, coursework, submissions):
        """Create sheet with statistical analysis"""
        ws = wb.create_sheet('Analysis')
        
        # Define styles
        title_font = Font(size=16, bold=True, color='2E5090')
        header_fill = PatternFill(start_color='2E5090', end_color='2E5090', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        alt_row_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        
        # Title
        ws['A1'] = 'Grade Analysis'
        ws['A1'].font = title_font
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Statistics table headers
        headers = ['Assignment', 'Average', 'Highest', 'Lowest', 'Submissions']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        row = 4
        total_students = len(students)
        for cw in coursework:
            cw_id = cw['id']
            cw_submissions = submissions.get(cw_id, [])
            max_pts = cw.get('maxPoints', 100)
            
            grades = [sub.get('assignedGrade') for sub in cw_submissions 
                     if sub.get('assignedGrade') is not None]
            
            # Count actual submissions (graded or turned in, not "NEW" state)
            actual_submissions = [sub for sub in cw_submissions 
                                 if sub.get('assignedGrade') is not None or 
                                    sub.get('state') in ['TURNED_IN', 'RETURNED']]
            
            # Alternating row colors
            if row % 2 == 0:
                row_fill = alt_row_fill
            else:
                row_fill = PatternFill()
            
            # Assignment name
            name_cell = ws.cell(row=row, column=1, value=cw['title'])
            name_cell.font = Font(bold=True, size=10)
            name_cell.fill = row_fill
            name_cell.border = border
            
            if grades:
                avg = sum(grades) / len(grades)
                avg_pct = (avg / max_pts * 100) if max_pts > 0 else 0
                
                avg_cell = ws.cell(row=row, column=2, value=f"{round(avg_pct, 1)}%")
                avg_cell.alignment = Alignment(horizontal='center', vertical='center')
                avg_cell.fill = row_fill
                avg_cell.border = border
                avg_cell.font = Font(size=10)
                
                high_cell = ws.cell(row=row, column=3, value=f"{max(grades)}/{max_pts}")
                high_cell.alignment = Alignment(horizontal='center', vertical='center')
                high_cell.fill = row_fill
                high_cell.border = border
                high_cell.font = Font(size=10, color='006400')
                
                low_cell = ws.cell(row=row, column=4, value=f"{min(grades)}/{max_pts}")
                low_cell.alignment = Alignment(horizontal='center', vertical='center')
                low_cell.fill = row_fill
                low_cell.border = border
                low_cell.font = Font(size=10, color='8B4513')
            
            sub_cell = ws.cell(row=row, column=5, value=f"{len(actual_submissions)}/{total_students}")
            sub_cell.alignment = Alignment(horizontal='center', vertical='center')
            sub_cell.fill = row_fill
            sub_cell.border = border
            sub_cell.font = Font(size=10)
            row += 1
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    # Skip merged cells
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                except:
                    pass
            if column_letter:
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
